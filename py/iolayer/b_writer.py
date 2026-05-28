"""
SlotPlanner Pro · S4
iolayer/b_writer.py — B 結果文件輸出

六大分頁：
  01_Summary        — 總覽 / RTP / 大獎
  02_RTP_By_Mode    — 各模式 RTP 與死局率
  03_Combo_Analysis — 連爆分佈（直方圖 + 累計）
  04_Symbol_Freq    — 符號出現頻率
  05_Discard_Stats  — 棄牌統計（HARD/SOFT 分開）
  06_Rule_Triggers  — 各腳本規則觸發次數與 RTP 貢獻
  (A_Config_Echo)   — A 設定檔關鍵參數回填（附錄）
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from core.schemas import AConfig

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ============================================================
# 樣式常數
# ============================================================
_ACCENT   = "5A3DB0"   # 薰衣草紫 header
_ACCENT2  = "7B5CC8"   # 次標題
_DEAD_CLR = "C0392B"   # 死局紅
_WIN_CLR  = "27AE60"   # 得獎綠
_WARN_CLR = "E67E22"   # 警告橘


def _hf(color: str = _ACCENT) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")

def _thin_border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws, cols: list[str], row: int = 1, color: str = _ACCENT):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill      = _hf(color)
        c.font      = _font(bold=True, color="FFFFFF")
        c.alignment = _center()


def _autofit(ws, min_width: int = 10, max_width: int = 40):
    for col in ws.columns:
        max_len = max_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value or "")
                max_len = min(max(len(v) + 2, min_width), max_width)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len


# ============================================================
# 公開入口
# ============================================================

def write_b_file(
    cfg: "AConfig",
    summary: dict,
    rules,
    output_dir: Path,
) -> Path:
    """
    產出 B 結果檔。

    :param cfg:        AConfig（用於回填設定）
    :param summary:    Collector.summary() 的回傳值
    :param rules:      cfg.puzzle_rules（已含 rtp_contribution）
    :param output_dir: 輸出資料夾
    :return:           寫出的 Path
    """
    if not _HAS_OPENPYXL:
        raise ImportError("請先安裝 openpyxl: pip install openpyxl")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{cfg.global_cfg.output_prefix}_{ts}.xlsx"
    out_path = output_dir / fname

    wb = Workbook()

    _write_summary(wb,       cfg, summary)
    _write_rtp_by_mode(wb,   cfg, summary)
    _write_combo_analysis(wb, summary)
    _write_symbol_freq(wb,    summary)
    _write_discard_stats(wb,  cfg, summary)
    _write_rule_triggers(wb,  rules, summary)
    _write_config_echo(wb,    cfg)

    wb.save(out_path)
    return out_path


# ============================================================
# 01_Summary
# ============================================================

def _write_summary(wb: Workbook, cfg: "AConfig", s: dict):
    ws = wb.active
    ws.title = "01_Summary"

    # 大標題
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "SlotPlanner Pro · 模擬結果總覽"
    c.font      = _font(bold=True, color="FFFFFF", size=14)
    c.fill      = _hf(_ACCENT)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    rows = [
        ("─── 基本資訊 ───",),
        ("模擬總局數",        f"{s['total_spins']:,}"),
        ("有效局數",          f"{s['valid_spins']:,}"),
        ("HARD 棄牌（風控）", f"{s['hard_discards']:,}"),
        ("SOFT 棄牌（體感）", f"{s['soft_discards']:,}"),
        ("死循環錯誤",        f"{s['loop_errors']:,}"),
        ("", ""),
        ("─── RTP ───",),
        ("理論 RTP",          f"{s['rtp_pct']:.4f}%"),
        ("總賠付（倍率加總）",f"{s['total_payout']:,.4f}"),
        ("中位數單局賠付",    f"{s['median_payout']:.6f}"),
        ("P95 單局賠付",      f"{s['p95_payout']:.6f}"),
        ("P99 單局賠付",      f"{s['p99_payout']:.6f}"),
        ("賠付類型",          cfg.global_cfg.pay_type.value),
        ("", ""),
        ("─── 死局 ───",),
        ("死局數",            f"{s['dead_spins']:,}"),
        ("死局率",            f"{s['dead_rate_pct']:.4f}%"),
        ("最長連續死局",      str(s['consecutive_dead_max'])),
        ("", ""),
        ("─── 大獎 ───",),
    ]
    # 大獎分層
    for label, cnt in s["big_win_counts"].items():
        pct = cnt / (s["valid_spins"] or 1) * 100
        rows.append((f"大獎 {label}", f"{cnt:,}  ({pct:.4f}%)"))

    rows += [
        ("", ""),
        ("─── 引擎資訊 ───",),
        ("Reel 數",           str(cfg.layout.reel_count)),
        ("符號數",            str(len(cfg.symbols))),
        ("腳本規則數",        str(len(cfg.puzzle_rules))),
        ("模式數",            str(len(cfg.modes))),
        ("起始模式",          cfg.global_cfg.starting_mode),
        ("隨機種子",          str(cfg.global_cfg.random_seed)),
    ]

    for r, row in enumerate(rows, 2):
        if len(row) == 1:
            # 分區標題
            ws.cell(row=r, column=1, value=row[0]).font = _font(bold=True, color=_ACCENT, size=11)
        else:
            ws.cell(row=r, column=1, value=row[0])
            ws.cell(row=r, column=2, value=row[1])

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 28


# ============================================================
# 02_RTP_By_Mode
# ============================================================

def _write_rtp_by_mode(wb: Workbook, cfg: "AConfig", s: dict):
    ws = wb.create_sheet("02_RTP_By_Mode")
    _write_header(ws, ["模式", "局數", "佔比%", "總賠付", "RTP%", "死局數", "死局率%"])

    mode_spins  = s["mode_spins"]
    mode_payout = s["mode_payout"]
    mode_dead   = s["mode_dead"]
    total_valid = s["valid_spins"] or 1

    for r, mode in enumerate(sorted(mode_spins.keys()), 2):
        spins  = mode_spins[mode]
        payout = mode_payout.get(mode, 0.0)
        dead   = mode_dead.get(mode, 0)
        rtp    = payout / spins * 100 if spins else 0
        dr     = dead / spins * 100 if spins else 0

        ws.cell(row=r, column=1, value=mode)
        ws.cell(row=r, column=2, value=spins)
        ws.cell(row=r, column=3, value=round(spins / total_valid * 100, 4))
        ws.cell(row=r, column=4, value=round(payout, 4))
        ws.cell(row=r, column=5, value=round(rtp, 4))
        ws.cell(row=r, column=6, value=dead)
        ws.cell(row=r, column=7, value=round(dr, 4))

    _autofit(ws)


# ============================================================
# 03_Combo_Analysis
# ============================================================

def _write_combo_analysis(wb: Workbook, s: dict):
    ws = wb.create_sheet("03_Combo_Analysis")
    _write_header(ws, ["連爆次數", "局數", "佔比%", "累計佔比%", "備註"])

    hist        = s["combo_hist"]
    total_valid = s["valid_spins"] or 1
    cum         = 0.0

    for r, steps in enumerate(sorted(hist.keys()), 2):
        cnt = hist[steps]
        pct = cnt / total_valid * 100
        cum += pct
        note = "死局" if steps == 0 else ""
        ws.cell(row=r, column=1, value=steps)
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=round(pct, 4))
        ws.cell(row=r, column=4, value=round(cum, 4))
        ws.cell(row=r, column=5, value=note)
        # 死局列標紅
        if steps == 0:
            for col in range(1, 6):
                ws.cell(row=r, column=col).font = _font(color=_DEAD_CLR)

    # 彙整列
    last_r = len(hist) + 2
    ws.cell(row=last_r + 1, column=1, value="中位數連爆").font = _font(bold=True)
    ws.cell(row=last_r + 1, column=2, value=s["median_combo"])

    _autofit(ws)


# ============================================================
# 04_Symbol_Freq
# ============================================================

def _write_symbol_freq(wb: Workbook, s: dict):
    ws = wb.create_sheet("04_Symbol_Freq")
    # v3.7 / #6:新增 4 個欄位(總賠付 / RTP 貢獻 / 賠付占比 / 每次出現平均賠付)
    _write_header(ws, [
        "Symbol_ID", "Display_Name", "Type",
        "出現次數", "出現佔比%",
        "總賠付", "RTP 貢獻%", "賠付占比%", "每次出現平均賠付",
    ])

    for r, (sid, info) in enumerate(s["symbol_freq"].items(), 2):
        ws.cell(row=r, column=1, value=sid)
        ws.cell(row=r, column=2, value=info["display_name"])
        ws.cell(row=r, column=3, value=info["sym_type"])
        ws.cell(row=r, column=4, value=info["count"])
        ws.cell(row=r, column=5, value=info["pct"])
        # v3.7 / #6
        ws.cell(row=r, column=6, value=info.get("payout", 0.0))
        ws.cell(row=r, column=7, value=info.get("rtp_contribution_pct", 0.0))
        ws.cell(row=r, column=8, value=info.get("payout_share_pct", 0.0))
        ws.cell(row=r, column=9, value=info.get("avg_payout_per_hit", 0.0))

    _autofit(ws)


# ============================================================
# 05_Discard_Stats
# ============================================================

def _write_discard_stats(wb: Workbook, cfg: "AConfig", s: dict):
    ws = wb.create_sheet("05_Discard_Stats")
    _write_header(ws, ["類型", "Rule_ID", "說明", "觸發次數", "觸發率%"])

    valid = s["valid_spins"] or 1
    r = 2

    # HARD 棄牌（總計）
    ws.cell(row=r, column=1, value="HARD（風控）")
    ws.cell(row=r, column=2, value="─ 合計 ─")
    ws.cell(row=r, column=4, value=s["hard_discards"])
    ws.cell(row=r, column=5, value=round(s["hard_discards"] / valid * 100, 4))
    for col in range(1, 6):
        ws.cell(row=r, column=col).font = _font(bold=True)
    r += 1

    # SOFT 棄牌（逐規則）
    soft_hits = s["soft_rule_hits"]
    soft_rules = {rule.rule_id: rule for rule in cfg.discard_rules
                  if rule.dtype.value == "SOFT"}

    if soft_hits:
        ws.cell(row=r, column=1, value="SOFT（體感）")
        ws.cell(row=r, column=2, value="─ 合計 ─")
        ws.cell(row=r, column=4, value=s["soft_discards"])
        ws.cell(row=r, column=5, value=round(s["soft_discards"] / valid * 100, 4))
        for col in range(1, 6):
            ws.cell(row=r, column=col).font = _font(bold=True)
        r += 1

        for rule_id, cnt in sorted(soft_hits.items(), key=lambda x: x[1], reverse=True):
            rule = soft_rules.get(rule_id)
            ws.cell(row=r, column=1, value="SOFT")
            ws.cell(row=r, column=2, value=rule_id)
            ws.cell(row=r, column=3, value=rule.reason_label if rule else "")
            ws.cell(row=r, column=4, value=cnt)
            ws.cell(row=r, column=5, value=round(cnt / valid * 100, 4))
            r += 1

    # 死局連續分佈
    r += 1
    ws.cell(row=r, column=1, value="連續死局分佈").font = _font(bold=True, color=_ACCENT)
    r += 1
    _write_header(ws, ["連續死局 Bucket", "次數", "佔比%"], row=r, color=_ACCENT2)
    r += 1
    for bucket, cnt in sorted(s["dead_bucket_counts"].items()):
        ws.cell(row=r, column=1, value=bucket)
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=round(cnt / valid * 100, 4))
        r += 1

    _autofit(ws)


# ============================================================
# 06_Rule_Triggers
# ============================================================

def _write_rule_triggers(wb: Workbook, rules, s: dict):
    ws = wb.create_sheet("06_Rule_Triggers")
    _write_header(ws, [
        "Rule_ID", "Priority", "Trigger", "Enabled",
        "觸發次數", "RTP 貢獻", "說明"
    ])

    for r, rule in enumerate(sorted(rules, key=lambda x: (x.priority, x.rule_id)), 2):
        ws.cell(row=r, column=1, value=rule.rule_id)
        ws.cell(row=r, column=2, value=rule.priority)
        ws.cell(row=r, column=3, value=rule.trigger.value)
        ws.cell(row=r, column=4, value="✓" if rule.enabled else "✗")
        ws.cell(row=r, column=5, value=rule.trigger_count)
        ws.cell(row=r, column=6, value=round(rule.rtp_contribution, 6))
        ws.cell(row=r, column=7, value=rule.description)

        if not rule.enabled:
            for col in range(1, 8):
                ws.cell(row=r, column=col).font = _font(color="999999")

    _autofit(ws)


# ============================================================
# A_Config_Echo（附錄）
# ============================================================

def _write_config_echo(wb: Workbook, cfg: "AConfig"):
    ws = wb.create_sheet("A_Config_Echo")

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = "A 設定檔關鍵參數（模擬當下的值）"
    c.font      = _font(bold=True, color="FFFFFF", size=12)
    c.fill      = _hf(_ACCENT2)
    c.alignment = _center()

    params = [
        ("simulation_count",    cfg.global_cfg.simulation_count),
        ("random_seed",         cfg.global_cfg.random_seed),
        ("pay_type",            cfg.global_cfg.pay_type.value),
        ("ways_direction",      cfg.global_cfg.ways_direction.value),
        ("cluster_min_size",    cfg.global_cfg.cluster_min_size),
        ("starting_mode",       cfg.global_cfg.starting_mode),
        ("max_chain_depth",     cfg.global_cfg.max_chain_depth),
        ("max_chain_per_rule",  cfg.global_cfg.max_chain_per_rule),
        ("big_win_thresholds",  str(cfg.global_cfg.big_win_thresholds)),
        ("dead_spin_buckets",   str(cfg.global_cfg.dead_spin_buckets)),
        ("reel_count",          cfg.layout.reel_count),
        ("symbol_count",        len(cfg.symbols)),
        ("payline_count",       len(cfg.paylines)),
        ("constraint_count",    len(cfg.constraints)),
        ("puzzle_rule_count",   len(cfg.puzzle_rules)),
        ("discard_rule_count",  len(cfg.discard_rules)),
        ("mode_count",          len(cfg.modes)),
    ]

    for r, (k, v) in enumerate(params, 2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=str(v))

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
